from flask import Flask, render_template,send_file, request, redirect, url_for, flash, session, jsonify
import mysql.connector
import random, string
import pandas as pd
import openpyxl
import smtplib
from fpdf import FPDF 
from email.message import EmailMessage
import os
from werkzeug.utils import secure_filename

app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = "college360"


UPLOAD_FOLDER = "static/uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# ================= DATABASE =================
def get_db():
    return mysql.connector.connect(
        host=os.getenv("DB_HOST"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASS"),
        database=os.getenv("DB_NAME"),
        port=int(os.getenv("DB_PORT"))
    )

# ================= UTILS =================
def generate_username(name, dept_code):
    return (
        name.split()[0].lower()
        + "_"
        + str(dept_code).lower()
        + str(random.randint(100, 999))
    )
def generate_password():
    return ''.join(random.choices(string.ascii_letters + string.digits, k=8))

# ================= EMAIL =================
import os
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

def send_email(to_email, username, password):
    # मजकूर साध्या फॉरमॅटमध्ये
    email_body = (
        f"You are appointed as Head of Department.\n\n"
        f"Username: {username}\n"
        f"Password: {password}\n\n"
        f"Login URL: https://college360-app.onrender.com/hod_login"
    )

    message = Mail(
        from_email='priyabhende11@gmail.com',
        to_emails=to_email,
        subject='HOD Login Credentials - College360',
        plain_text_content=email_body
    )

    try:
        # तुमची खरी API Key (शुद्ध केलेली)
        # की च्या शेवटी कोणतेही जादा अक्षरे (उदा. cI) नसावेत याची खात्री करा
        api_key = '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'
        
        sg = SendGridAPIClient(api_key)
        response = sg.send(message)
        
        print(f"✅ Email sent successfully! Status Code: {response.status_code}")
        
    except Exception as e:
        # जर तरीही 401 आला, तर नवीन API Key जनरेट करणे उत्तम
        print(f"❌ SendGrid Error: {str(e)}")
# ================= HOME =================
@app.route("/")
def index():
    return render_template("index.html")

principal_username = "APMatale"
principal_password = "1234"
# ================= PRINCIPAL LOGIN =================
@app.route("/principal_login", methods=["GET", "POST"])
def principal_login():
    global principal_username, principal_password

    if request.method == "POST":
        if (request.form["username"] == principal_username and 
            request.form["password"] == principal_password):

            session.clear()
            session["role"] = "principal"
            session["principal"] = True
            return redirect(url_for("principal_dashboard"))

        flash("Invalid Login")

    return render_template("principal_login.html")

@app.route("/forgot_password", methods=["GET", "POST"])
def forgot_password():
    global principal_username, principal_password
    if request.method == "POST":
        principal_username = request.form["new_username"]
        principal_password = request.form["new_password"]

        return redirect(url_for("principal_login"))
        flash("Username and Password Updated Successfully!")
    return render_template("forgot_password.html")

@app.route("/principal_dashboard")
def principal_dashboard():
    if session.get("role") != "principal":
        return redirect(url_for("principal_login"))
    return render_template("principal_dashboard.html")


# ================= ADD DEPARTMENT =================
@app.route("/add_dept", methods=["GET", "POST"])
def add_dept():
    if "principal" not in session:
        return redirect(url_for("principal_login"))

    if request.method == "POST":
        db = get_db()
        cur = db.cursor()
        cur.execute(
            "INSERT INTO departments (dept_name, dept_code) VALUES (%s,%s)",
            (request.form["dept_name"], request.form["dept_code"])
        )
        db.commit()
        db.close()
        flash("Department Added Successfully")
        return redirect(url_for("principal_dashboard"))

    return render_template("add_dept.html")


@app.route("/employee_add", methods=["GET", "POST"])
def employee_add():

    # 🔐 Login check
    if session.get("role") not in ["principal", "hod"]:
        return redirect(url_for("index"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    # 📌 Load departments
    if session["role"] == "principal":
        cur.execute("SELECT dept_id, dept_name FROM departments")
        departments = cur.fetchall()
    else:
        cur.execute("""
            SELECT dept_id, dept_name
            FROM departments
            WHERE dept_id=%s
        """, (session["dept_id"],))
        departments = cur.fetchall()

    # 📌 Handle POST
    if request.method == "POST":

        # 🧠 dept logic
        if session["role"] == "hod":
            dept_id = session["dept_id"]
        else:
            dept_id = request.form["dept_id"]

        cur.execute("""
    INSERT INTO staff
    (staff_name, dob, contact, email, gender,
     qualification, semester, subject, subject_code,
     address, dept_id, role)
    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'staff')
""", (
    request.form["name"],
    request.form["dob"],
    request.form["contact"],
    request.form["email"],
    request.form["gender"],
    request.form["qualification"],
    request.form["semester"],
    request.form["subject"],
    request.form["subject_code"],   # ✅ MUST exist
    request.form["address"],
    dept_id
))

        db.commit()
        flash("Staff added successfully")

        return redirect(url_for("employee_add"))

    db.close()
    return render_template(
        "employee_add.html",
        departments=departments
    )



# ================= ASSIGN HOD PAGE =================
@app.route("/assign_hod")
def assign_hod_page():
    if "principal" not in session:
        return redirect(url_for("principal_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT dept_id, dept_name FROM departments")
    departments = cur.fetchall()
    db.close()

    return render_template("assign_hod.html", departments=departments)

# ================= GET STAFF BY DEPARTMENT =================
@app.route("/get_staff_by_department/<int:dept_id>")
def get_staff_by_department(dept_id):
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT staff_id, staff_name, email
        FROM staff
        WHERE dept_id=%s AND role='staff'
    """, (dept_id,))

    staff = cur.fetchall()
    db.close()
    return jsonify(staff)

# ================= ASSIGN HOD ACTION =================
@app.route("/assign_hod_action", methods=["POST"])
def assign_hod_action():
    if "principal" not in session:
        return jsonify({"success": False})

    staff_id = request.json["staff_id"]

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ✅ 1. get selected staff
    cur.execute("SELECT * FROM staff WHERE staff_id=%s", (staff_id,))
    staff = cur.fetchone()

    if not staff:
        db.close()
        return jsonify({"success": False, "msg": "Staff not found"})

    # ✅ 2. get department code
    cur.execute(
        "SELECT dept_code FROM departments WHERE dept_id=%s",
        (staff["dept_id"],)
    )
    dept = cur.fetchone()
    dept_code = dept["dept_code"]

    # ✅ 3. generate credentials
    username = generate_username(staff["staff_name"], dept_code)
    password = generate_password()

    # ✅ 4. remove previous HOD of same dept
    cur.execute("""
        UPDATE staff
        SET role='staff'
        WHERE dept_id=%s AND role='hod'
    """, (staff["dept_id"],))

    # ✅ 5. assign new HOD
    cur.execute("""
        UPDATE staff
        SET role='hod', username=%s, password=%s
        WHERE staff_id=%s
    """, (username, password, staff_id))

    db.commit()
    db.close()

    return jsonify({
        "success": True,
        "username": username,
        "password": password,
        "email": staff["email"]
    })
    
@app.route("/view_staff/<int:staff_id>")
def view_staff(staff_id):
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT s.*, d.dept_name
        FROM staff s
        JOIN departments d ON s.dept_id = d.dept_id
        WHERE s.staff_id=%s
    """, (staff_id,))

    staff = cur.fetchone()
    db.close()

    return render_template("view_staff.html", staff=staff)



@app.route("/edit_staff/<int:staff_id>", methods=["GET", "POST"])
def edit_staff(staff_id):
    db = get_db()
    cur = db.cursor(dictionary=True)

    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        contact = request.form["contact"]
        address = request.form["address"]
        qualification = request.form["qualification"]
        role = request.form["role"]

        photo_file = request.files["photo"]
        photo_name = None

        if photo_file and photo_file.filename != "":
            photo_name = secure_filename(photo_file.filename)
            photo_file.save(os.path.join(app.config["UPLOAD_FOLDER"], photo_name))

            cur.execute("""
                UPDATE staff SET
                staff_name=%s, email=%s, contact=%s,
                address=%s, qualification=%s, role=%s,
                photo=%s
                WHERE staff_id=%s
            """, (name, email, contact, address, qualification, role, photo_name, staff_id))
        else:
            cur.execute("""
                UPDATE staff SET
                staff_name=%s, email=%s, contact=%s,
                address=%s, qualification=%s, role=%s
                WHERE staff_id=%s
            """, (name, email, contact, address, qualification, role, staff_id))

        db.commit()
        db.close()
        flash("Staff updated successfully")
        return redirect(url_for("view_hod_staff"))

    cur.execute("SELECT * FROM staff WHERE staff_id=%s", (staff_id,))
    staff = cur.fetchone()
    db.close()

    return render_template("edit_staff.html", staff=staff)


# ================= HOD LOGIN =================
@app.route("/hod_login", methods=["GET", "POST"])
def hod_login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        db = get_db()
        cur = db.cursor(dictionary=True)

        cur.execute("""
            SELECT staff_id, dept_id
            FROM staff
            WHERE username=%s AND password=%s AND role='hod'
        """, (username, password))

        hod = cur.fetchone()
        db.close()

        if hod:
            # session.clear() 
            session["hod"] = True  # 🔴 IMPORTANT
            session["role"] = "hod"         # 🔴 THIS WAS MISSING
            session["hod_id"] = hod["staff_id"]
            session["dept_id"] = hod["dept_id"]

            return redirect(url_for("hod_dashboard"))

        flash("Invalid Username or Password")

    return render_template("hod_login.html")


@app.route("/hod_dashboard")
def hod_dashboard():
    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute(
        "SELECT dept_name FROM departments WHERE dept_id=%s",
        (session["dept_id"],)
    )
    dept = cur.fetchone()
    db.close()

    return render_template(
        "hod_dashboard.html",
        dept_name=dept["dept_name"]
    )



# ================= LOGOUT =================
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))

@app.route("/mess")
def mess():
    return render_template("Mess.html")

@app.route("/mou")
def mou():
    return render_template("MOU.html")

@app.route("/hod/mou/add", methods=["GET", "POST"])
def hod_mou_add():
    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    edit_mou = None

    # ================= INSERT / UPDATE =================
    if request.method == "POST":
        mou_id = request.form.get("mou_id")

        if mou_id:   # ✏️ UPDATE
            cur.execute("""
                UPDATE mou
                SET title=%s,
                    organization=%s,
                    start_date=%s,
                    end_date=%s,
                    description=%s
                WHERE mou_id=%s AND dept_id=%s AND is_deleted=0
            """, (
                request.form["title"],
                request.form["organization"],
                request.form["start_date"],
                request.form["end_date"],
                request.form["description"],
                mou_id,
                session["dept_id"]
            ))
            flash("MOU Updated Successfully")

        else:        # ➕ INSERT
            cur.execute("""
                INSERT INTO mou
                (dept_id, title, organization, start_date, end_date,
                 description, status, is_deleted, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,'Active',0,%s)
            """, (
                session["dept_id"],
                request.form["title"],
                request.form["organization"],
                request.form["start_date"],
                request.form["end_date"],
                request.form["description"],
                session["hod_id"]
            ))
            flash("MOU Added Successfully")

        db.commit()
        return redirect(url_for("hod_mou_add"))

    # ================= VIEW LIST =================
    cur.execute("""
        SELECT *
        FROM mou
        WHERE dept_id=%s AND is_deleted=0
        ORDER BY start_date DESC
    """, (session["dept_id"],))

    mou_list = cur.fetchall()
    db.close()

    return render_template(
        "hod_mou_add.html",
        mou_list=mou_list,
        edit_mou=edit_mou
    )


@app.route("/hod/mou/edit/<int:mou_id>")
def hod_mou_edit(mou_id):
    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    # 👉 Selected MOU (for edit)
    cur.execute("""
        SELECT *
        FROM mou
        WHERE mou_id=%s AND dept_id=%s AND is_deleted=0
    """, (mou_id, session["dept_id"]))

    edit_mou = cur.fetchone()

    # 👉 All MOU list
    cur.execute("""
        SELECT *
        FROM mou
        WHERE dept_id=%s AND is_deleted=0
        ORDER BY start_date DESC
    """, (session["dept_id"],))

    mou_list = cur.fetchall()
    db.close()

    return render_template(
        "hod_mou_add.html",
        mou_list=mou_list,
        edit_mou=edit_mou
    )

@app.route("/hod/mou")
def hod_mou_view():
    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT *
        FROM MOU
        WHERE dept_id=%s
        ORDER BY start_date DESC
    """, (session["dept_id"],))

    data = cur.fetchall()
    db.close()

    return render_template("hod_mou_view.html", data=data)

@app.route("/hod/mou/delete/<int:mou_id>")
def hod_mou_delete(mou_id):
    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    db = get_db()
    cur = db.cursor()

    cur.execute("""
        UPDATE mou
        SET status='Inactive', is_deleted=1
        WHERE mou_id=%s AND dept_id=%s
    """, (mou_id, session["dept_id"]))

    db.commit()
    db.close()

    flash("MOU Expired Successfully")
    return redirect(url_for("hod_mou_add"))


@app.route("/hod/mou/status/<int:mou_id>")
def hod_mou_status(mou_id):
    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT status FROM mou
        WHERE mou_id=%s AND dept_id=%s AND is_deleted=0
    """, (mou_id, session["dept_id"]))

    mou = cur.fetchone()

    if mou:
        new_status = "Inactive" if mou["status"] == "Active" else "Active"
        cur.execute("""
            UPDATE mou
            SET status=%s
            WHERE mou_id=%s AND dept_id=%s
        """, (new_status, mou_id, session["dept_id"]))
        db.commit()

    db.close()
    return redirect(url_for("hod_mou_add"))

@app.route("/principal/mou", methods=["GET", "POST"])
def principal_mou():
    if session.get("role") != "principal":
        return redirect(url_for("principal_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    # Department dropdown
    cur.execute("SELECT dept_id, dept_name FROM departments ORDER BY dept_name")
    departments = cur.fetchall()

    selected_dept = request.form.get("dept_id")

    query = """
        SELECT 
            d.dept_name,
            m.title,
            m.organization,
            m.start_date,
            m.end_date,
            m.status
        FROM mou m
        JOIN departments d ON m.dept_id = d.dept_id
        WHERE m.status='Active' AND m.is_deleted=0
    """
    params = []

    if selected_dept and selected_dept != "all":
        query += " AND m.dept_id=%s"
        params.append(selected_dept)

    query += " ORDER BY d.dept_name, m.start_date DESC"

    cur.execute(query, params)
    data = cur.fetchall()

    db.close()

    return render_template(
        "principal_mou_view.html",
        data=data,
        departments=departments,
        selected_dept=selected_dept
    )


@app.route("/activity/k8")
def activity_k8_redirect():
    role = session.get("role")

    if role == "hod":
        return redirect(url_for("add_k8"))
    elif role == "principal":
        return redirect(url_for("view_k8"))
    else:
        return redirect(url_for("index"))

# ➕ HOD ADD K8
from werkzeug.utils import secure_filename

@app.route("/hod/k8/add", methods=["GET", "POST"])
def add_k8():

    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    app.config["UPLOAD_FOLDER"] = "static/uploads"
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ✅ ALWAYS fetch department name first (for GET + POST)
    cur.execute(
        "SELECT dept_name FROM departments WHERE dept_id=%s",
        (session["dept_id"],)
    )
    dept = cur.fetchone()
    dept_name = dept["dept_name"] if dept else ""

    if request.method == "POST":

        photos = request.files.getlist("photo")
        filenames = []

        for photo in photos:
            if photo and photo.filename != "":
                import uuid
                filename = str(uuid.uuid4()) + "_" + secure_filename(photo.filename)
                photo.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
                filenames.append(filename)

        photo_string = ",".join(filenames) if filenames else None

        cur.execute("""
            INSERT INTO k8_industrial_visit
            (academic_year, dept_id, visit_date,
             industry_name, coordinator_name,
             beneficiaries, relevance,
             mapping_level, semester, photo)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            request.form["academic_year"],
            session["dept_id"],  # 🔒 dept auto from session
            request.form["visit_date"],
            request.form["industry_name"],
            request.form["coordinator_name"],
            request.form["beneficiaries"],
            request.form["relevance"],
            request.form["mapping_level"],
            request.form["semester"],
            photo_string
        ))

        db.commit()
        flash("Industrial Visit Added Successfully")

        return redirect(url_for("add_k8"))

    db.close()

    return render_template(
        "add_k8.html",
        dept_name=dept_name,
        back_url=url_for("act"),
        view_url=url_for("view_k8")
    )
# 👁 VIEW K8 (HOD + PRINCIPAL)
@app.route("/k8/view")
def view_k8():

    role = session.get("role")
    if role not in ["hod", "principal"]:
        return redirect(url_for("index"))

    selected_year = request.args.get("academic_year", "").strip()
    selected_dept = request.args.get("dept_id", "").strip()

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ✅ Academic years dropdown data
    cur.execute("SELECT DISTINCT academic_year FROM k8_industrial_visit ORDER BY academic_year DESC")
    academic_years = [row["academic_year"] for row in cur.fetchall()]

    # ✅ Department dropdown
    if role == "principal":
        cur.execute("SELECT dept_id, dept_name FROM departments")
        departments = cur.fetchall()
    else:
        departments = []

    query = """
        SELECT k8.*, d.dept_name
        FROM k8_industrial_visit k8
        JOIN departments d ON k8.dept_id = d.dept_id
        WHERE 1=1
    """
    params = []

    # 🔒 HOD restriction
    if role == "hod":
        query += " AND k8.dept_id = %s"
        params.append(session["dept_id"])

    # 🎯 Principal selected dept
    if selected_dept:
        query += " AND k8.dept_id = %s"
        params.append(selected_dept)

    # 🎯 Academic year filter
    if selected_year:
        query += " AND k8.academic_year = %s"
        params.append(selected_year)

    cur.execute(query, params)
    records = cur.fetchall()
    db.close()

    return render_template(
        "view_k8.html",
        records=records,
        academic_years=academic_years,
        departments=departments,
        selected_year=selected_year,
        selected_dept=selected_dept,
        role=role,
        back_url=url_for("act") if role == "principal" else url_for("add_k8")
    )



@app.route("/activity/k9")
def activity_k9_redirect():
    role = session.get("role")

    if role == "hod":
        return redirect(url_for("hod_k9_add"))
    elif role == "principal":
        return redirect(url_for("principal_k9_view"))
    else:
        return redirect(url_for("index"))

@app.route("/hod/k9/add", methods=["GET", "POST"])
def hod_k9_add():

    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    app.config['UPLOAD_FOLDER'] = 'static/uploads'
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ✅ ALWAYS fetch department name (GET + POST)
    cur.execute(
        "SELECT dept_name FROM departments WHERE dept_id=%s",
        (session["dept_id"],)
    )
    dept = cur.fetchone()
    dept_name = dept["dept_name"] if dept else ""

    if request.method == 'POST':

        photos = request.files.getlist("photo")
        filenames = []

        import uuid
        from werkzeug.utils import secure_filename

        for photo in photos:
            if photo and photo.filename != "":
                filename = str(uuid.uuid4()) + "_" + secure_filename(photo.filename)
                photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                filenames.append(filename)

        photo_string = ",".join(filenames) if filenames else None

        cur.execute("""
            INSERT INTO k9_expert_lecture 
            (dept_id, academic_year, programme, 
             expert_name, designation, organization, 
             email, lecture_date, topic, year_sem, 
             coordinator, students, po_pso, photo, created_by) 
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            session["dept_id"],
            request.form["academic_year"],
            dept_name,  # ✅ programme auto from department
            request.form["expert_name"],
            request.form["designation"],
            request.form["organization"],
            request.form["email"],
            request.form["lecture_date"],
            request.form["topic"],
            request.form["year_sem"],
            request.form["coordinator"],
            request.form["students"],
            request.form["po_pso"],
            photo_string,
            session["hod_id"]
        ))

        db.commit()
        flash("Expert Lecture Added Successfully")

        return redirect(url_for("hod_k9_add"))

    db.close()

    return render_template(
        "add_k9.html",
        dept_name=dept_name
    )
@app.route("/hod/k9/view")
def hod_k9_view():

    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    year = request.args.get("academic_year")  # ✅ same name

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = "SELECT * FROM k9_expert_lecture WHERE dept_id=%s"
    params = [session["dept_id"]]

    if year:
        query += " AND academic_year=%s"
        params.append(year)

    cur.execute(query, params)
    data = cur.fetchall()

    # Dropdown values
    cur.execute("""
        SELECT DISTINCT academic_year
        FROM k9_expert_lecture
        WHERE dept_id=%s
        ORDER BY academic_year DESC
    """, (session["dept_id"],))

    academic_years = [r["academic_year"] for r in cur.fetchall()]

    db.close()

    return render_template(
        "view_k9.html",
        data=data,
        academic_years=academic_years,
        selected_year=year,
        programme=data[0]["programme"] if data else ""
    )
@app.route("/principal/k9/view")
def principal_k9_view():

    if session.get("role") != "principal":
        return redirect(url_for("principal_login"))

    year = request.args.get("academic_year")
    dept_id = request.args.get("dept_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    query = """
        SELECT k.*, d.dept_name
        FROM k9_expert_lecture k
        JOIN departments d ON k.dept_id = d.dept_id
        WHERE 1=1
    """

    params = []

    if year:
        query += " AND k.academic_year=%s"
        params.append(year)

    if dept_id:
        query += " AND k.dept_id=%s"
        params.append(dept_id)

    cur.execute(query, params)
    data = cur.fetchall()

    # Academic Years
    cur.execute("""
        SELECT DISTINCT academic_year
        FROM k9_expert_lecture
        ORDER BY academic_year DESC
    """)
    academic_years = [r["academic_year"] for r in cur.fetchall()]

    # Departments
    cur.execute("""
        SELECT dept_id, dept_name
        FROM departments
        ORDER BY dept_name
    """)
    departments = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        "view_k9.html",
        data=data,
        academic_years=academic_years,
        departments=departments,
        selected_year=year,
        selected_dept=dept_id,
        role="principal"
    )
@app.route("/activity/k10")
def activity_k10():
    role = session.get("role")

    if role == "hod":
        return redirect(url_for("add_k10"))

    elif role == "principal":
        return redirect(url_for("view_k10"))

    else:
        return redirect(url_for("index"))
    
@app.route("/tpo")
def tpo():
    return render_template("TPO.html")

@app.route("/add_k10", methods=["GET", "POST"])
def add_k10():

    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ✅ ALWAYS fetch department name (GET + POST safe)
    cur.execute(
        "SELECT dept_name FROM departments WHERE dept_id=%s",
        (session["dept_id"],)
    )
    dept = cur.fetchone()
    dept_name = dept["dept_name"] if dept else ""

    if request.method == "POST":

        dept_id = session["dept_id"]
        user_id = session["hod_id"]

        # ================= PART A =================
        if request.form.get("save_part") == "A":

            cur.execute("""
                INSERT INTO k10_industry_placement
                (dept_id, academic_year, programme, visit_date,
                 industry_name, industry_address, industry_type,
                 students_attended, students_placed, salary, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                dept_id,
                request.form["a_academic_year"],
                dept_name,   # ✅ AUTO programme
                request.form["a_date"],
                request.form["industry_name"],
                request.form["industry_address"],
                request.form["industry_type"],
                request.form["students_attended"],
                request.form["students_placed"],
                request.form["salary"],
                user_id
            ))

            flash("PART-A saved successfully")

        # ================= PART B =================
        if request.form.get("save_part") == "B":

            cur.execute("""
                INSERT INTO k10_student_placement
                (dept_id, academic_year, programme, student_name,
                 placement_nature, organization, remark, created_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                dept_id,
                request.form["b_academic_year"],
                dept_name,   # ✅ AUTO programme
                request.form["student_name"],
                request.form["placement_nature"],
                request.form["organization"],
                request.form["remark"],
                user_id
            ))

            flash("PART-B saved successfully")

        db.commit()
        return redirect(url_for("add_k10"))

    db.close()

    return render_template(
        "add_k10.html",
        dept_name=dept_name
    )
@app.route("/view_k10")
def view_k10():

    if "role" not in session:
        return redirect(url_for("login"))

    role = session["role"]
    year = request.args.get("academic_year")
    dept = request.args.get("dept_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ================= YEARS =================
    if role == "hod":
        cur.execute("""
            SELECT DISTINCT academic_year
            FROM k10_industry_placement
            WHERE dept_id=%s
            ORDER BY academic_year DESC
        """, (session["dept_id"],))
    else:
        cur.execute("""
            SELECT DISTINCT academic_year
            FROM k10_industry_placement
            ORDER BY academic_year DESC
        """)

    years = [r["academic_year"] for r in cur.fetchall()]

    # ================= DEPARTMENTS (ONLY PRINCIPAL) =================
    departments = []
    if role == "principal":
        cur.execute("SELECT dept_id, dept_name FROM departments ORDER BY dept_name")
        departments = cur.fetchall()

    # ================= PART-A =================
    query_a = """
        SELECT k.*, d.dept_name
        FROM k10_industry_placement k
        JOIN departments d ON k.dept_id=d.dept_id
        WHERE 1=1
    """
    params_a = []

    if role == "hod":
        query_a += " AND k.dept_id=%s"
        params_a.append(session["dept_id"])

    if year:
        query_a += " AND k.academic_year=%s"
        params_a.append(year)

    if role == "principal" and dept:
        query_a += " AND k.dept_id=%s"
        params_a.append(dept)

    cur.execute(query_a, params_a)
    part_a = cur.fetchall()

    # ================= PART-B =================
    query_b = """
        SELECT k.*, d.dept_name
        FROM k10_student_placement k
        JOIN departments d ON k.dept_id=d.dept_id
        WHERE 1=1
    """
    params_b = []

    if role == "hod":
        query_b += " AND k.dept_id=%s"
        params_b.append(session["dept_id"])

    if year:
        query_b += " AND k.academic_year=%s"
        params_b.append(year)

    if role == "principal" and dept:
        query_b += " AND k.dept_id=%s"
        params_b.append(dept)

    cur.execute(query_b, params_b)
    part_b = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        "view_k10.html",
        part_a=part_a,
        part_b=part_b,
        years=years,
        departments=departments,
        selected_year=year,
        selected_dept=dept,
        role=role,
        back_url=url_for("act")
    )

    
# ---------- K13 MAIN REDIRECT ----------
@app.route("/activity/k13")
def activity_k13():
    role = session.get("role")

    if role == "hod":
        return redirect(url_for("add_k13"))
    elif role == "principal":
        return redirect(url_for("view_k13"))
    else:
        return redirect(url_for("index"))


# ---------- ADD K13 ----------
@app.route("/add_k13", methods=["GET", "POST"])
def add_k13():

    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    app.config["UPLOAD_FOLDER"] = "static/uploads"
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ✅ ALWAYS fetch department name (GET + POST safe)
    cur.execute(
        "SELECT dept_name FROM departments WHERE dept_id=%s",
        (session["dept_id"],)
    )
    dept = cur.fetchone()
    dept_name = dept["dept_name"] if dept else ""

    if request.method == "POST":

        photos = request.files.getlist("photo")
        filenames = []

        import uuid
        from werkzeug.utils import secure_filename

        for photo in photos:
            if photo and photo.filename != "":
                filename = str(uuid.uuid4()) + "_" + secure_filename(photo.filename)
                photo.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
                filenames.append(filename)

        photo_string = ",".join(filenames) if filenames else None

        organizer = request.form["organizer"]
        organizer_other = request.form.get("organizer_other")

        if organizer != "Other":
            organizer_other = None

        cur.execute("""
            INSERT INTO k13_co_curricular
(academic_year, dept_id,dept_name,
 activity_date, activity_type,
 organizer, organizer_other,
 student_name, enrollment_no, award_prize,
 participants, po_pso, photo, created_by)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            request.form["academic_year"],
            session["dept_id"],
            dept_name,  # ✅ AUTO programme
            request.form["activity_date"],
            request.form["activity_type"],
            organizer,
            organizer_other,
            request.form["student_name"],
            request.form["enrollment_no"],
            request.form["award_prize"],
            request.form["participants"],
            request.form["po_pso"],
            photo_string,
            session["hod_id"]
        ))

        db.commit()
        flash("K13 Activity Added Successfully")

        return redirect(url_for("add_k13"))

    db.close()

    return render_template(
        "add_k13.html",
        dept_name=dept_name
    )
@app.route("/delete_photo/<table_name>", methods=["POST"])
def delete_photo(table_name):

    data = request.get_json()
    record_id = data.get("id")
    photo_name = data.get("photo")

    db = get_db()
    cur = db.cursor()

    # Step 1 – Get current photo column
    cur.execute(f"SELECT photo FROM {table_name} WHERE id=%s", (record_id,))
    row = cur.fetchone()

    if row:
        photos = row[0].split(",")
        photos = [p.strip() for p in photos if p.strip() != photo_name]

        updated_photos = ",".join(photos)

        cur.execute(f"""
            UPDATE {table_name}
            SET photo=%s
            WHERE id=%s
        """, (updated_photos, record_id))

        db.commit()

    cur.close()
    db.close()

    return {"success": True}
# ---------- VIEW K13 ----------
@app.route("/view_k13")
def view_k13():

    if "role" not in session:
        return redirect(url_for("login"))

    role = session["role"]
    year = request.args.get("academic_year")
    dept_id = request.args.get("dept_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ================= BASE QUERY =================
    query = """
        SELECT k.*, d.dept_name
        FROM k13_co_curricular k
        JOIN departments d ON k.dept_id = d.dept_id
        WHERE 1=1
    """
    params = []

    # ================= ROLE FILTER =================
    if role == "hod":
        query += " AND k.dept_id=%s"
        params.append(session["dept_id"])

    # ================= YEAR FILTER =================
    if year:
        query += " AND k.academic_year=%s"
        params.append(year)

    # ================= DEPT FILTER (PRINCIPAL ONLY) =================
    if role == "principal" and dept_id:
        query += " AND k.dept_id=%s"
        params.append(dept_id)

    cur.execute(query, params)
    records = cur.fetchall()

    # ================= YEARS DROPDOWN =================
    if role == "hod":
        cur.execute("""
            SELECT DISTINCT academic_year
            FROM k13_co_curricular
            WHERE dept_id=%s
            ORDER BY academic_year DESC
        """, (session["dept_id"],))
    else:
        cur.execute("""
            SELECT DISTINCT academic_year
            FROM k13_co_curricular
            ORDER BY academic_year DESC
        """)

    academic_years = [r["academic_year"] for r in cur.fetchall()]

    # ================= DEPARTMENTS DROPDOWN =================
    departments = []
    if role == "principal":
        cur.execute("SELECT dept_id, dept_name FROM departments ORDER BY dept_name")
        departments = cur.fetchall()

    cur.close()
    db.close()

    # ================= BACK BUTTON =================
    if role == "hod":
        back_url = url_for("act")              # Activity.html
    else:
        back_url = url_for("principal_dashboard")

    return render_template(
        "view_k13.html",
        records=records,
        academic_years=academic_years,
        departments=departments,
        selected_year=year,
        selected_dept=dept_id,
        role=role,
        back_url=back_url,
        institute_name="Government Polytechnic Yavatmal",
        programme="Diploma Engineering"
    )


@app.route("/act")
def act():
    role = session.get("role")

    if role == "hod":
        back_url = url_for("hod_dashboard")
    elif role == "principal":
        back_url = url_for("principal_dashboard")
    else:
        back_url = url_for("index")

    return render_template("Activity.html", back_url=back_url)

@app.route("/activity/k14")
def activity_k14():
    role = session.get("role")

    if role == "hod":
        return redirect(url_for("add_k14"))
    elif role == "principal":
        return redirect(url_for("view_k14"))
    else:
        return redirect(url_for("index"))
    
app.config['UPLOAD_FOLDER'] = 'static/k14_photos'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)  
 
@app.route("/add_k14", methods=["GET", "POST"])
def add_k14():

    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    app.config["UPLOAD_FOLDER"] = "static/uploads"
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

    db = get_db()
    cur = db.cursor(dictionary=True)

    # Department name fetch
    cur.execute(
        "SELECT dept_name FROM departments WHERE dept_id=%s",
        (session["dept_id"],)
    )
    dept = cur.fetchone()
    dept_name = dept["dept_name"] if dept else ""

    if request.method == "POST":

        academic_year = request.form.get("academic_year")
        activity_date = request.form.get("activity_date")
        activity_type = request.form.get("activity_type")
        activity_details = request.form.get("activity_details")
        organizing = request.form.get("organizing")
        organizing_other = request.form.get("organizing_other")
        participant_name = request.form.get("participant_name")
        enrollment_no = request.form.get("enrollment_no")
        award_prize = request.form.get("award_prize")
        relevance_po = request.form.get("relevance_po")

        # PHOTO UPLOAD
        photos = request.files.getlist("photo")
        filenames = []

        import uuid
        from werkzeug.utils import secure_filename

        for photo in photos:
            if photo and photo.filename != "":
                filename = str(uuid.uuid4()) + "_" + secure_filename(photo.filename)

                path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                photo.save(path)

                filenames.append(filename)

        photo_string = ",".join(filenames) if filenames else None

        # INSERT QUERY
        cur.execute("""
            INSERT INTO k14_extra_curricular
            (academic_year, dept_id, dept_name,
            activity_date, activity_type, activity_details,
            organizing, organizing_other,
            participant_name, enrollment_no,
            award_prize, relevance_po,
            activity_photo)

            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            academic_year,
            session["dept_id"],
            dept_name,
            activity_date,
            activity_type,
            activity_details,
            organizing,
            organizing_other,
            participant_name,
            enrollment_no,
            award_prize,
            relevance_po,
            photo_string
        ))

        db.commit()
        flash("K14 Activity Saved Successfully")

        return redirect(url_for("add_k14"))

    db.close()

    return render_template(
        "add_k14.html",
        dept_name=dept_name
    )
app.secret_key = "college360_secret_key"
@app.route('/view_k14')
def view_k14():

    if 'role' not in session:
        return redirect(url_for('login'))

    role = session['role']
    year = request.args.get('academic_year')
    dept = request.args.get('dept_id')

    db = get_db()
    cur = db.cursor(dictionary=True)
    

    # ================= BASE QUERY =================
    query = """
        SELECT k.*, d.dept_name
        FROM k14_extra_curricular k
        JOIN departments d ON k.dept_id = d.dept_id
        WHERE 1=1
    """
    params = []

    # ================= ROLE FILTER =================
    if role == "hod":
        query += " AND k.dept_id=%s"
        params.append(session["dept_id"])

    # ================= YEAR FILTER =================
    if year:
        query += " AND k.academic_year=%s"
        params.append(year)

    # ================= DEPT FILTER (ONLY PRINCIPAL) =================
    if role == "principal" and dept:
        query += " AND k.dept_id=%s"
        params.append(dept)

    # ================= EXECUTE =================
    cur.execute(query, params)
    data = cur.fetchall()

    # ================= YEARS =================
    if role == "hod":
        cur.execute("""
            SELECT DISTINCT academic_year
            FROM k14_extra_curricular
            WHERE dept_id=%s
            ORDER BY academic_year DESC
        """, (session["dept_id"],))
    else:
        cur.execute("""
            SELECT DISTINCT academic_year
            FROM k14_extra_curricular
            ORDER BY academic_year DESC
        """)

    years = [r["academic_year"] for r in cur.fetchall()]

    # ================= DEPARTMENTS (PRINCIPAL ONLY) =================
    departments = []
    if role == "principal":
        cur.execute("SELECT dept_id, dept_name FROM departments ORDER BY dept_name")
        departments = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        'view_k14.html',
        data=data,
        years=years,
        departments=departments,
        selected_year=year,
        selected_dept=dept,
        role=role,
        back_url=url_for("act")
    )

@app.route("/activity/k11")
def activity_k11():
    role = session.get("role")

    if role == "hod":
        return redirect(url_for("add_k11"))
    elif role == "principal":
        return redirect(url_for("view_k11"))
    else:
        return redirect(url_for("login"))

@app.route("/add_k11", methods=["GET", "POST"])
def add_k11():

    if session.get("role") != "hod":
        return redirect(url_for("hod_login"))

    app.config['UPLOAD_FOLDER'] = 'static/uploads'
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ✅ ALWAYS fetch department name (GET + POST safe)
    cur.execute(
        "SELECT dept_name FROM departments WHERE dept_id=%s",
        (session["dept_id"],)
    )
    dept = cur.fetchone()
    dept_name = dept["dept_name"] if dept else ""

    if request.method == 'POST':

        photos = request.files.getlist("photo")
        filenames = []

        import uuid
        from werkzeug.utils import secure_filename

        for photo in photos:
            if photo and photo.filename != "":
                filename = str(uuid.uuid4()) + "_" + secure_filename(photo.filename)
                photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                filenames.append(filename)

        photo_string = ",".join(filenames) if filenames else None

        cur.execute("""
            INSERT INTO k11_faculty_training
            (academic_year, dept_id, programme,
             name_designation, training_details, mode, duration,
             training_date, organizing_body, organizing_institute,
             photo, created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            request.form["academic_year"],
            session["dept_id"],
            dept_name,   # ✅ AUTO programme
            request.form["name_designation"],
            request.form["training_details"],
            request.form["mode"],
            request.form["duration"],
            request.form["date"],
            request.form["organizing_body"],
            request.form["organizing_institute"],
            photo_string,
            session["hod_id"]
        ))

        db.commit()
        flash("K11 Training record added successfully")

        return redirect(url_for("add_k11"))

    db.close()

    return render_template(
        "add_k11.html",
        dept_name=dept_name
    )
@app.route("/view_k11")
def view_k11():

    if "role" not in session:
        return redirect(url_for("login"))
    
       # ✅ AUTO

    role = session["role"]
    year = request.args.get("academic_year")
    dept_id = request.args.get("dept_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ================= BASE QUERY =================
    query = """
        SELECT k.*, d.dept_name
        FROM k11_faculty_training k
        JOIN departments d ON k.dept_id = d.dept_id
        WHERE 1=1
    """
    params = []

    # ================= ROLE FILTER =================
    if role == "hod":
        query += " AND k.dept_id=%s"
        params.append(session["dept_id"])

    # ================= YEAR FILTER =================
    if year:
        query += " AND k.academic_year=%s"
        params.append(year)

    # ================= DEPT FILTER (PRINCIPAL ONLY) =================
    if role == "principal" and dept_id:
        query += " AND k.dept_id=%s"
        params.append(dept_id)

    cur.execute(query, params)
    records = cur.fetchall()

    # ================= YEARS =================
    if role == "hod":
        cur.execute("""
            SELECT DISTINCT academic_year
            FROM k11_faculty_training
            WHERE dept_id=%s
            ORDER BY academic_year DESC
        """, (session["dept_id"],))
    else:
        cur.execute("""
            SELECT DISTINCT academic_year
            FROM k11_faculty_training
            ORDER BY academic_year DESC
        """)

    academic_years = [r["academic_year"] for r in cur.fetchall()]

    # ================= DEPARTMENTS =================
    departments = []
    if role == "principal":
        cur.execute("SELECT dept_id, dept_name FROM departments ORDER BY dept_name")
        departments = cur.fetchall()

    cur.close()
    db.close()

    # ================= BACK URL =================
    if role == "hod":
        back_url = url_for("act")          # Activity.html
    else:
        back_url = url_for("act")

    return render_template(
        "view_k11.html",
        records=records,
        academic_years=academic_years,
        departments=departments,
        selected_year=year,
        selected_dept=dept_id,
        role=role,
        back_url=back_url,
        institute_name="Government Polytechnic Yavatmal",
        
    )

@app.route("/testing")
def testing():
    return render_template("testing_add.html")

# @app.route("/testing/back")
# def testing_back():
#     role = session.get("role")

#     if role == "hod":
#         return redirect(url_for("hod_dashboard"))

#     elif role == "principal":
#         return redirect(url_for("principal_dashboard"))

#     return redirect(url_for("index"))


@app.route("/activity/testing")
def testing_redirect():
    role = session.get("role")

    if role == "hod":
        return redirect(url_for("testing_add"))
    elif role == "principal":
        return redirect(url_for("testing_view"))
    else:
        return redirect(url_for("index"))
    
@app.route("/hod/testing/add", methods=["GET","POST"])
def testing_add():

    if session.get("role") != "hod":
        return redirect(url_for("principal_dashboard"))

    if request.method == "POST":

        db = get_db()
        cur = db.cursor()

        dept_id = session.get("dept_id")   # 🔒 Safe get

        cur.execute("""
            INSERT INTO Testing
            (academic_year, dept_id, tester_name, tester_contact,
             location, test_date, expense, remarks)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            request.form.get("academic_year"),
            dept_id,
            request.form.get("tester_name"),
            request.form.get("tester_contact"),
            request.form.get("location"),
            request.form.get("test_date"),
            request.form.get("expense") or 0,
            request.form.get("remarks")
        ))

        db.commit()
        cur.close()
        db.close()

        flash("Lab Testing Added Successfully")

        return redirect(url_for("testing_add"))

    return render_template(
        "testing_add.html",
        back_url=url_for("hod_dashboard"),
        view_url=url_for("testing_view")
    )


@app.route("/testing/view")
def testing_view():

    role = session.get("role")

    if role not in ["hod", "principal"]:
        return redirect(url_for("index"))

    year = request.args.get("academic_year")
    selected_dept = request.args.get("dept_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # 🔹 Department dropdown
    cur.execute("SELECT dept_id, dept_name FROM departments")
    departments = cur.fetchall()

    # 🔹 Academic year dropdown
    cur.execute("SELECT DISTINCT academic_year FROM testing ORDER BY academic_year DESC")
    academic_years = [r["academic_year"] for r in cur.fetchall()]

    # 🔹 Base query
    query = """
        SELECT lt.*, d.dept_name
        FROM Testing lt
        JOIN departments d ON lt.dept_id = d.dept_id
        WHERE 1=1
    """
    params = []

    # 🔒 HOD restriction
    if role == "hod":
        query += " AND lt.dept_id=%s"
        params.append(session.get("dept_id"))

    # 🔽 Principal department filter
    if role == "principal" and selected_dept:
        query += " AND lt.dept_id=%s"
        params.append(selected_dept)

    # 🔽 Academic year filter
    if year:
        query += " AND lt.academic_year=%s"
        params.append(year)

    query += " ORDER BY lt.test_date DESC"

    cur.execute(query, tuple(params))
    records = cur.fetchall()

    cur.close()
    db.close()

    back_url = url_for("hod_dashboard") if role == "hod" else url_for("principal_dashboard")

    return render_template(
        "testing_view.html",
        records=records,
        academic_years=academic_years,
        departments=departments,
        selected_year=year,
        selected_dept=selected_dept,
        back_url=back_url
    )


@app.route("/hostel")
def hostel():
    return render_template("Hostel.html")

@app.route("/department/co")
def dept_co():
    return render_template("dept_co.html")

@app.route("/department/ee")
def dept_ee():
    return render_template("dept_ee.html")

@app.route("/department/ej")
def dept_ej():
    return render_template("dept_ej.html")

@app.route("/department/me")
def dept_me():
    return render_template("dept_me.html")

@app.route("/department/civil")
def dept_civil():
    return render_template("dept_civil.html")


@app.route("/view_hod_staff")
def view_hod_staff():
    if "principal" not in session:
        return redirect(url_for("principal_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT d.dept_name, s.staff_id, s.staff_name, s.email, s.role
        FROM staff s
        JOIN departments d ON s.dept_id = d.dept_id
        ORDER BY d.dept_name
    """)
    data = cur.fetchall()
    db.close()

    return render_template("view_hod_staff.html", data=data)

@app.route("/resend_hod_email/<int:staff_id>")
def resend_hod_email(staff_id):
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT staff_name, email, username, password
        FROM staff WHERE staff_id=%s AND role='hod'
    """, (staff_id,))

    hod = cur.fetchone()
    db.close()

    if hod:
        send_email(hod["email"], hod["username"], hod["password"])
        flash("Email sent successfully to HOD")

    return redirect(url_for("view_hod_staff"))

@app.route("/send_staff_email/<int:staff_id>")
def send_staff_email(staff_id):
    if "principal" not in session:
        return redirect(url_for("principal_login"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    # get staff
    cur.execute("SELECT * FROM staff WHERE staff_id=%s", (staff_id,))
    staff = cur.fetchone()

    if not staff:
        flash("Staff not found")
        return redirect(url_for("view_hod_staff"))

    # get dept code
    cur.execute(
        "SELECT dept_code FROM departments WHERE dept_id=%s",
        (staff["dept_id"],)
    )
    dept = cur.fetchone()
    dept_code = dept["dept_code"]

    # generate credentials if not exists
    if not staff["username"] or not staff["password"]:
        username = generate_username(staff["staff_name"], dept_code)
        password = generate_password()

        cur.execute("""
            UPDATE staff
            SET username=%s, password=%s
            WHERE staff_id=%s
        """, (username, password, staff_id))
        db.commit()
    else:
        username = staff["username"]
        password = staff["password"]

    db.close()

    send_email(staff["email"], username, password)
    flash("Login credentials sent successfully")

    return redirect(url_for("view_hod_staff"))

@app.route("/exam/add")
@app.route("/activity/k7")
def activity_k7():
    return redirect(url_for("bulk_upload"))


@app.route("/registration_dashboard")
def registration_dashboard():
    # dept wise dashboard
    return render_template(
        "registration_dashboard.html",
        dept_name=session.get("dept_name")
    )
@app.route("/students")
def students():

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT id, scheme_name FROM schemes ORDER BY id DESC")
    schemes = cur.fetchall()

    cur.close()
    db.close()

    return render_template("student_reg.html", schemes=schemes)
@app.route("/view_students")
def view_students():
    db = get_db()
    cur = db.cursor(dictionary=True)

    # 🔵 GET FILTER VALUE
    scheme = request.args.get("scheme")

    # 🔵 LOAD STUDENTS
    if scheme:
        cur.execute("SELECT * FROM students_data WHERE scheme=%s ORDER BY student_id DESC", (scheme,))
    else:
        cur.execute("SELECT * FROM students_data ORDER BY student_id DESC")

    students = cur.fetchall()

    # 🔵 LOAD ONLY USED SCHEMES
    cur.execute("""
        SELECT DISTINCT scheme 
        FROM students_data 
        WHERE scheme IS NOT NULL AND scheme != ''
        ORDER BY scheme
    """)
    schemes = cur.fetchall()

    cur.close()
    db.close()

    return render_template("view_students.html", students=students, schemes=schemes)
@app.route("/add_student", methods=["POST"])
def add_student():
    data = request.form
    dept_id  = request.args.get("dept_id")
    db = get_db()
    cur = db.cursor()

    cur.execute("""
        INSERT INTO students_data
        (academic_year, diploma_year,semester,scheme, enrollment_no, student_name, dept,
         contact, email, father_name, mother_name, father_email, father_contact)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data["academic_year"],
        data["diploma_year"],
        data["semester"],
        data["scheme"],
        data["enrollment_no"],
        data["student_name"],
        data["dept"],
        data["contact"],
        data["email"],
        data["father_name"],
        data["mother_name"],
        data["father_email"],
        data["father_contact"]
    ))

    db.commit()
    cur.close()
    db.close()

    flash("Student added successfully")
    return redirect(url_for("students"))
@app.route("/download_student_template")
def download_student_template():
    df = pd.DataFrame(columns=[
        "academic_year",
        "diploma_year",
        "semester",
        "scheme",
        "enrollment_no",
        "student_name",
        "dept",
        "contact",
        "email",
        "father_name",
        "mother_name",
        "father_email",
        "father_contact"
    ])

    file_path = "student_template.xlsx"
    df.to_excel(file_path, index=False)

    return send_file(file_path, as_attachment=True)

@app.route("/bulk_upload_students", methods=["POST"])
def bulk_upload_students():
    file = request.files["file"]

    df = pd.read_excel(file)
    df.columns = df.columns.str.strip().str.lower()

    # Replace NaN with empty string
    df = df.fillna("")

    db = get_db()
    cur = db.cursor()

    for _, row in df.iterrows():
        cur.execute("""
            INSERT INTO students_data
            (academic_year, diploma_year,semester,scheme, enrollment_no, student_name, dept,
             contact, email, father_name, mother_name, father_email, father_contact)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            row["academic_year"],
            row["diploma_year"],
            row["semester"],
            row["scheme"],
            row["enrollment_no"],
            row["student_name"],
            row["dept"],
            row["contact"],
            row["email"],
            row["father_name"],
            row["mother_name"],
            row["father_email"],
            row["father_contact"]
        ))

    db.commit()
    cur.close()
    db.close()

    flash("Bulk upload completed successfully")
    return redirect(url_for("students"))

@app.route("/courses")
def courses():

    dept_id = session.get("dept_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # Dept name fetch
    cur.execute("SELECT dept_name FROM departments WHERE dept_id=%s", (dept_id,))
    dept = cur.fetchone()

    # Scheme fetch
    cur.execute("SELECT id, scheme_name FROM schemes")
    schemes = cur.fetchall()

    cur.close()
    db.close()

    return render_template("courses.html",
                           dept_id=dept_id,
                           dept_name=dept["dept_name"],
                           schemes=schemes)


@app.route("/view")
def course_view():

    dept_id = session.get("dept_id")
    selected_scheme = request.args.get("scheme_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # Only schemes used in courses of this dept
    cur.execute("""
        SELECT DISTINCT s.id, s.scheme_name
        FROM schemes s
        JOIN courses c ON s.id = c.scheme_id
        WHERE c.dept_id = %s
    """, (dept_id,))
    schemes = cur.fetchall()

    query = """
        SELECT c.*, d.dept_name, s.scheme_name,
               s.head1, s.head2, s.head3, s.head4, s.head5
        FROM courses c
        JOIN departments d ON c.dept_id = d.dept_id
        JOIN schemes s ON c.scheme_id = s.id
        WHERE c.dept_id = %s
    """
    params = [dept_id]

    if selected_scheme:
        query += " AND c.scheme_id = %s"
        params.append(selected_scheme)

    cur.execute(query, tuple(params))
    courses = cur.fetchall()

    cur.close()
    db.close()

    return render_template("course_view.html",
                           courses=courses,
                           schemes=schemes,
                           selected_scheme=selected_scheme)
@app.route("/save", methods=["POST"])
def save_course():

    db = get_db()
    cur = db.cursor()

    data = request.form

    # 🔥 Safe integer converter
    def safe_int(value):
        if value is None or value.strip() == "":
            return None   # NULL store hoil
        return int(value)

    head1 = safe_int(data.get("head1_marks"))
    head2 = safe_int(data.get("head2_marks"))
    head3 = safe_int(data.get("head3_marks"))
    head4 = safe_int(data.get("head4_marks"))
    head5 = safe_int(data.get("head5_marks"))

    cur.execute("""
        INSERT INTO courses
        (course_code, course_title, year, dept_id, semester, scheme_id,
         head1_marks, head2_marks, head3_marks, head4_marks, head5_marks)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data["course_code"],
        data["course_title"],
        data["year"],
        session.get("dept_id"),
        data["semester"],
        data["scheme_id"],
        head1,
        head2,
        head3,
        head4,
        head5
    ))

    db.commit()
    cur.close()
    db.close()

    return redirect("/courses?success=1")
@app.route("/add_cos")
def add_cos():

    dept_id = session.get("dept_id")
    dept_name = session.get("dept_name")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # Only courses of that dept
    cur.execute("SELECT course_code, course_title FROM courses WHERE dept_id=%s", (dept_id,))
    courses = cur.fetchall()

    # Schemes dropdown
    cur.execute("SELECT id, scheme_name FROM schemes")
    schemes = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        "add_cos.html",
        courses=courses,
        schemes=schemes,
        dept_id=dept_id,
        dept_name=dept_name
    )

@app.route("/co/save", methods=["POST"])
def save_cos():

    year = request.form["year"]
    semester = request.form["semester"]
    dept_id = request.form["dept_id"]
    scheme_id = request.form["scheme_id"]

    course_code = request.form["course_code"]
    course_title = request.form["course_title"]

    co1 = request.form["co1"]
    co2 = request.form.get("co2")
    co3 = request.form.get("co3")
    co4 = request.form.get("co4")
    co5 = request.form.get("co5")

    db = get_db()
    cur = db.cursor()

    cur.execute("""
        INSERT INTO course_outcomes
        (year, semester, dept_id, scheme_id,
         course_code, course_title,
         co1, co2, co3, co4, co5)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        year, semester, dept_id, scheme_id,
        course_code, course_title,
        co1, co2, co3, co4, co5
    ))

    db.commit()
    cur.close()
    db.close()

    return redirect(url_for("add_cos", success=1))

@app.route("/view_cos")
def view_cos():

    selected_scheme = request.args.get("scheme_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT id, scheme_name FROM schemes")
    schemes = cur.fetchall()

    query = "SELECT * FROM course_outcomes WHERE 1=1"
    params = []

    if selected_scheme:
        query += " AND scheme_id=%s"
        params.append(selected_scheme)

    cur.execute(query, tuple(params))
    cos = cur.fetchall()

    cur.close()
    db.close()

    return render_template("view_cos.html",
                           cos=cos,
                           schemes=schemes,
                           selected_scheme=selected_scheme)
    
@app.route("/get_courses/<scheme_id>")
def get_courses(scheme_id):

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT course_code, course_title
        FROM courses
        WHERE scheme_id = %s
    """, (scheme_id,))

    courses = cur.fetchall()

    cur.close()
    db.close()

    return jsonify(courses)

@app.route("/fa_th_analysis")
def fa_th_analysis():

    semester = request.args.get("semester")
    scheme_id = request.args.get("scheme_id")
    dept_id = session.get("dept_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # Departments
    cur.execute("SELECT * FROM departments")
    departments = cur.fetchall()

    # Schemes (ALL from schemes table)
    cur.execute("SELECT id, scheme_name FROM schemes")
    schemes = cur.fetchall()

    # Courses FILTERED BY SCHEME (if selected)
    if scheme_id:
        cur.execute("SELECT * FROM courses WHERE scheme_id=%s", (scheme_id,))
    else:
        cur.execute("SELECT * FROM courses")
    courses = cur.fetchall()

    # Staff
    cur.execute("SELECT * FROM staff")
    staff = cur.fetchall()

    students = []

    if semester and dept_id:
        cur.execute("""
            SELECT enrollment_no, student_name
            FROM students_data
            WHERE semester=%s
            ORDER BY enrollment_no
        """, (semester,))
        students = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        "add_k7A.html",
        departments=departments,
        schemes=schemes,
        selected_scheme=scheme_id,
        courses=courses,
        staff=staff,
        students=students,
        selected_sem=semester
    )
    
@app.route("/fa_th/save", methods=["POST"])
def save_fa_th():

    d = request.form

    academic_year = d.get("academic_year")

    # -------- SAFE INT FUNCTION --------
    def safe_int(value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    dept_id = safe_int(d.get("dept_id"))
    semester = safe_int(d.get("semester"))
    scheme_id = d.get("scheme_id")

    if not scheme_id:
        return "Scheme is required"

    scheme_id = int(scheme_id)
    staff_id = safe_int(d.get("staff_id"))

    exam = d.get("exam")
    course_code = d.get("course_code")
    course_title = d.get("course_title")

    enrollment_list = d.getlist("enrollment_no[]")

    def to_int(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return 0

    db = get_db()
    cur = db.cursor()

    for i in range(len(enrollment_list)):

        q1a = to_int(d.getlist("q1a[]")[i])
        q1b = to_int(d.getlist("q1b[]")[i])
        q1c = to_int(d.getlist("q1c[]")[i])
        q1d = to_int(d.getlist("q1d[]")[i])
        q1e = to_int(d.getlist("q1e[]")[i])
        q1f = to_int(d.getlist("q1f[]")[i])
        q1g = to_int(d.getlist("q1g[]")[i])
        q1h = to_int(d.getlist("q1h[]")[i])

        q2a = to_int(d.getlist("q2a[]")[i])
        q2b = to_int(d.getlist("q2b[]")[i])
        q2c = to_int(d.getlist("q2c[]")[i])
        q2d = to_int(d.getlist("q2d[]")[i])
        q2e = to_int(d.getlist("q2e[]")[i])
        q2f = to_int(d.getlist("q2f[]")[i])
        q2g = to_int(d.getlist("q2g[]")[i])
        q2h = to_int(d.getlist("q2h[]")[i])

        q3a = to_int(d.getlist("q3a[]")[i])
        q3b = to_int(d.getlist("q3b[]")[i])
        q3c = to_int(d.getlist("q3c[]")[i])

        total = to_int(d.getlist("total[]")[i])

        cur.execute("""
            INSERT INTO fa_th_analysis
            (academic_year, dept_id, semester, scheme_id, exam_type,
             course_code, course_title, staff_id, enrollment_no,
             q1a,q1b,q1c,q1d,q1e,q1f,q1g,q1h,
             q2a,q2b,q2c,q2d,q2e,q2f,q2g,q2h,
             q3a,q3b,q3c,total)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s)
        """, (
            academic_year, dept_id, semester, scheme_id, exam,
            course_code, course_title, staff_id,
            enrollment_list[i],
            q1a,q1b,q1c,q1d,q1e,q1f,q1g,q1h,
            q2a,q2b,q2c,q2d,q2e,q2f,q2g,q2h,
            q3a,q3b,q3c,total
        ))

    db.commit()
    cur.close()
    db.close()

    return redirect(url_for("fa_th_analysis",
                            semester=semester,
                            scheme_id=scheme_id))
@app.route("/view_fa_th")
def view_fa_th():

    db = get_db()
    cur = db.cursor(dictionary=True)

    selected_year = request.args.get("academic_year")
    selected_sem = request.args.get("semester")
    selected_exam = request.args.get("exam")
    selected_course = request.args.get("course_code")
    selected_scheme = request.args.get("scheme_id")

    # 🔽 Dropdown data (DISTINCT values)
    cur.execute("SELECT DISTINCT academic_year FROM fa_th_analysis")
    years = cur.fetchall()

    cur.execute("SELECT DISTINCT semester FROM fa_th_analysis")
    semesters = cur.fetchall()

    cur.execute("SELECT DISTINCT exam_type FROM fa_th_analysis")
    exams = cur.fetchall()

    cur.execute("SELECT DISTINCT course_code FROM fa_th_analysis")
    courses = cur.fetchall()

    cur.execute("SELECT id, scheme_name FROM schemes")
    schemes = cur.fetchall()

    # 🔽 Main Query
    query = """
        SELECT *
        FROM fa_th_analysis
        WHERE 1=1
    """
    params = []

    if selected_year:
        query += " AND academic_year=%s"
        params.append(selected_year)

    if selected_sem:
        query += " AND semester=%s"
        params.append(selected_sem)

    if selected_exam:
        query += " AND exam_type=%s"
        params.append(selected_exam)

    if selected_course:
        query += " AND course_code=%s"
        params.append(selected_course)

    if selected_scheme:
        query += " AND scheme_id=%s"
        params.append(selected_scheme)

    cur.execute(query, tuple(params))
    records = cur.fetchall()

    cur.close()
    db.close()

    return render_template("fa_th_view.html",
                           records=records,
                           years=years,
                           semesters=semesters,
                           exams=exams,
                           courses=courses,
                           schemes=schemes,
                           selected_year=selected_year,
                           selected_sem=selected_sem,
                           selected_exam=selected_exam,
                           selected_course=selected_course,
                           selected_scheme=selected_scheme)
@app.route("/generate")
def generate():

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT DISTINCT academic_year FROM fa_th_analysis")
    years = cur.fetchall()

    cur.execute("""
        SELECT DISTINCT d.dept_id, d.dept_code
        FROM fa_th_analysis f
        JOIN departments d ON f.dept_id = d.dept_id
    """)
    departments = cur.fetchall()

    cur.execute("SELECT DISTINCT course_code, course_title FROM fa_th_analysis")
    courses = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        "generate_k7A.html",
        years=years,
        departments=departments,
        courses=courses,
        totals={},        # 🔴 VERY IMPORTANT
        dept_id="",
        course_code="",
        academic_year=""
    )

@app.route("/attaintment_k7", methods=["GET", "POST"])
def attaintment_k7():

    dept_id = request.values.get("dept_id")
    course_code = request.values.get("course_code")
    academic_year = request.values.get("academic_year")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ✅ Correct student count
    cur.execute("""
        SELECT COUNT(DISTINCT enrollment_no) AS total_students
        FROM fa_th_analysis
        WHERE dept_id=%s AND course_code=%s AND academic_year=%s
    """, (dept_id, course_code, academic_year))

    student_data = cur.fetchone()
    total_students = student_data["total_students"] if student_data else 0


    # ✅ Question totals
    cur.execute("""
        SELECT
          COUNT(CASE WHEN q1a>0 THEN 1 END)*2 AS Q1a,
          COUNT(CASE WHEN q1b>0 THEN 1 END)*2 AS Q1b,
          COUNT(CASE WHEN q1c>0 THEN 1 END)*2 AS Q1c,
          COUNT(CASE WHEN q1d>0 THEN 1 END)*2 AS Q1d,
          COUNT(CASE WHEN q1e>0 THEN 1 END)*2 AS Q1e,
          COUNT(CASE WHEN q1f>0 THEN 1 END)*2 AS Q1f,
          COUNT(CASE WHEN q1g>0 THEN 1 END)*2 AS Q1g,
          COUNT(CASE WHEN q1h>0 THEN 1 END)*2 AS Q1h,

          COUNT(CASE WHEN q2a>0 THEN 1 END)*4 AS Q2a,
          COUNT(CASE WHEN q2b>0 THEN 1 END)*4 AS Q2b,
          COUNT(CASE WHEN q2c>0 THEN 1 END)*4 AS Q2c,
          COUNT(CASE WHEN q2d>0 THEN 1 END)*4 AS Q2d,
          COUNT(CASE WHEN q2e>0 THEN 1 END)*4 AS Q2e,
          COUNT(CASE WHEN q2f>0 THEN 1 END)*4 AS Q2f,
          COUNT(CASE WHEN q2g>0 THEN 1 END)*4 AS Q2g,
          COUNT(CASE WHEN q2h>0 THEN 1 END)*4 AS Q2h,

          COUNT(CASE WHEN q3a>0 THEN 1 END)*6 AS Q3a,
          COUNT(CASE WHEN q3b>0 THEN 1 END)*6 AS Q3b,
          COUNT(CASE WHEN q3c>0 THEN 1 END)*6 AS Q3c
        FROM fa_th_analysis
        WHERE dept_id=%s AND course_code=%s AND academic_year=%s
    """, (dept_id, course_code, academic_year))

    totals = cur.fetchone() or {}

    # Save CO mapping
    if request.method == "POST":

        questions = request.form.getlist("question[]")
        cos = request.form.getlist("co[]")

        cur.execute("""
            DELETE FROM attainment_k7
            WHERE dept_id=%s AND course_code=%s AND academic_year=%s
        """, (dept_id, course_code, academic_year))

        for q, co in zip(questions, cos):

            cur.execute("""
                INSERT INTO attainment_k7
                (dept_id, course_code, academic_year, question, co_no)
                VALUES (%s,%s,%s,%s,%s)
            """, (dept_id, course_code, academic_year, q, co))

        db.commit()

    cur.close()
    db.close()

    return render_template(
        "attaintment_k7.html",
        totals=totals,
        total_students=total_students,
        dept_id=dept_id,
        course_code=course_code,
        academic_year=academic_year
    )

@app.route("/view_attaintment_k7")
def view_attaintment_k7():

    selected_year = request.args.get("academic_year")

    db = get_db()
    cur = db.cursor(dictionary=True)

    # 🔹 dropdown years (only from attainment table – SAFE)
    cur.execute("""
        SELECT DISTINCT TRIM(academic_year) AS academic_year
        FROM attainment_k7
        ORDER BY academic_year DESC
    """)
    years = cur.fetchall()

    records = []

    if selected_year:
        cur.execute("""
            SELECT
                TRIM(a.academic_year) AS academic_year,
                a.dept_id,
                a.course_code,
                a.question,
                a.co_no
            FROM attainment_k7 a
            WHERE TRIM(a.academic_year) = %s
            ORDER BY question
        """, (selected_year,))
        records = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        "view_attaintment_k7.html",
        years=years,
        records=records,
        selected_year=selected_year
    )

    
@app.route("/k7/save", methods=["POST"])
def save_k7():

    d = request.form
    db = get_db()
    cur = db.cursor()

    questions = d.getlist("question")
    totals = d.getlist("total")
    cos = d.getlist("co")

    for q, t, co in zip(questions, totals, cos):
        cur.execute("""
            INSERT INTO attainment_k7
            (academic_year, dept_id, course_code, exam_type,
             question, total_marks, co_no)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            d["academic_year"],
            d["dept_id"],
            d["course_code"],
            "FA-TH",
            q,
            t,
            co
        ))

    db.commit()
    cur.close()
    db.close()

    # flash("✅ K7 Attainment Saved Successfully")
    return redirect(url_for("generate"))


@app.route("/api/student/<enroll>")
def get_student(enroll):
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT student_name, dept
        FROM students_data
        WHERE enrollment_no=%s
    """, (enroll,))

    student = cur.fetchone()
    cur.close()
    db.close()

    return student if student else {}

import json

@app.route("/k7b", methods=["GET", "POST"])
def k7b():

    if session.get("role") == "principal":
        return redirect(url_for("activity"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ----------------------
    # Static Data
    # ----------------------

    cur.execute("""
        SELECT DISTINCT s.id, s.scheme_name
        FROM schemes s
        JOIN gazette_1 g ON s.id = g.scheme_id
    """)
    schemes = cur.fetchall()

    cur.execute("SELECT DISTINCT academic_year FROM gazette_1 ORDER BY academic_year DESC")
    years = cur.fetchall()

    cur.execute("SELECT DISTINCT semester FROM gazette_1 ORDER BY semester")
    semesters = cur.fetchall()

    cur.execute("SELECT dept_id, dept_name FROM departments")
    departments = cur.fetchall()

    courses = []
    show_marks = False
    stats = {}

    scheme_id = year = semester = dept_id = subject_name = None

    if request.method == "POST":

        scheme_id = request.form.get("scheme_id")
        year = request.form.get("year")
        semester = request.form.get("semester")
        dept_id = request.form.get("dept_id")
        subject_name = request.form.get("course_code")

        selected_dept = dept_id if session.get("role") == "principal" else session.get("dept_id")

        # ----------------------
        # Load subjects
        # ----------------------

        if year and semester and selected_dept and scheme_id:

            cur.execute("""
                SELECT DISTINCT subject_name
                FROM gazette_1
                WHERE academic_year=%s 
                AND semester=%s 
                AND dept_id=%s 
                AND scheme_id=%s
            """, (year, semester, selected_dept, scheme_id))

            courses = cur.fetchall()

        # ----------------------
        # Load marks_json
        # ----------------------

        if year and semester and selected_dept and scheme_id:

            query = """
                SELECT subject_name, marks_json
                FROM gazette_1
                WHERE academic_year=%s 
                AND semester=%s 
                AND dept_id=%s 
                AND scheme_id=%s
            """

            params = [year, semester, selected_dept, scheme_id]

            if subject_name:
                query += " AND subject_name=%s"
                params.append(subject_name)

            cur.execute(query, tuple(params))
            all_rows = cur.fetchall()

            show_marks = True
            stats = {}

            # ----------------------
            # Extract marks
            # ----------------------

            for row in all_rows:

                subj = row["subject_name"]

                try:
                    marks_dict = json.loads(row["marks_json"])
                except:
                    continue

                if subj not in stats:
                    stats[subj] = {}

                for head, mark in marks_dict.items():

                    try:
                        mark = float(mark)
                    except:
                        continue      # skip AB / NA / empty

                    if mark <= 0:
                        continue      # absent / detain

                    if head not in stats[subj]:
                        stats[subj][head] = {"marks_list": []}

                    stats[subj][head]["marks_list"].append(mark)

            # ----------------------
            # Calculate statistics
            # ----------------------

            for subj, heads in stats.items():

                for head, data in heads.items():

                    marks_list = data["marks_list"]

                    if not marks_list:
                        stats[subj][head] = {
                            "min": "-",
                            "max": "-",
                            "appeared": 0,
                            "passed": 0,
                            "pass_percent": 0,
                            "above60": 0,
                            "na": True
                        }
                        continue

                    appeared = len(marks_list)

                    min_marks = min(marks_list)
                    max_marks = max(marks_list)

                    pass_mark = max_marks * 0.4
                    above60_mark = max_marks * 0.6

                    passed = len([m for m in marks_list if m >= pass_mark])
                    above60 = len([m for m in marks_list if m >= above60_mark])

                    stats[subj][head] = {
                        "min": min_marks,
                        "max": max_marks,
                        "appeared": appeared,
                        "passed": passed,
                        "pass_percent": round((passed / appeared) * 100, 2) if appeared else 0,
                        "above60": round((above60 / appeared) * 100, 2) if appeared else 0,
                        "na": False
                    }

    cur.close()
    db.close()

    return render_template(
        "k7B.html",
        schemes=schemes,
        years=years,
        semesters=semesters,
        departments=departments,
        courses=courses,
        show_marks=show_marks,
        stats=stats,
        scheme_id=scheme_id,
        year=year,
        semester=semester,
        dept_id=dept_id,
        course_code=subject_name
    )

    if session.get("role") == "principal":
        return redirect(url_for("activity"))

    db = get_db()
    cur = db.cursor(dictionary=True)

    # ----------------------
    # Static Data
    # ----------------------

    cur.execute("""
        SELECT DISTINCT s.id, s.scheme_name
        FROM schemes s
        JOIN gazette_1 g ON s.id = g.scheme_id
    """)
    schemes = cur.fetchall()

    cur.execute("SELECT DISTINCT academic_year FROM gazette_1 ORDER BY academic_year DESC")
    years = cur.fetchall()

    cur.execute("SELECT DISTINCT semester FROM gazette_1 ORDER BY semester")
    semesters = cur.fetchall()

    cur.execute("SELECT dept_id, dept_name FROM departments")
    departments = cur.fetchall()

    courses = []
    show_marks = False
    stats = {}

    scheme_id = year = semester = dept_id = subject_name = None

    if request.method == "POST":

        scheme_id = request.form.get("scheme_id")
        year = request.form.get("year")
        semester = request.form.get("semester")
        dept_id = request.form.get("dept_id")
        subject_name = request.form.get("course_code")

        selected_dept = dept_id if session.get("role") == "principal" else session.get("dept_id")

        # ----------------------
        # Load subjects
        # ----------------------

        if year and semester and selected_dept and scheme_id:

            cur.execute("""
                SELECT DISTINCT subject_name
                FROM gazette_1
                WHERE academic_year=%s 
                AND semester=%s 
                AND dept_id=%s 
                AND scheme_id=%s
            """, (year, semester, selected_dept, scheme_id))

            courses = cur.fetchall()

        # ----------------------
        # Load marks_json
        # ----------------------

        if year and semester and selected_dept and scheme_id:

            query = """
                SELECT subject_name, marks_json
                FROM gazette_1
                WHERE academic_year=%s 
                AND semester=%s 
                AND dept_id=%s 
                AND scheme_id=%s
            """

            params = [year, semester, selected_dept, scheme_id]

            if subject_name:
                query += " AND subject_name=%s"
                params.append(subject_name)

            cur.execute(query, tuple(params))
            all_rows = cur.fetchall()

            show_marks = True
            stats = {}

            # ----------------------
            # Extract marks
            # ----------------------

            for row in all_rows:

                subj = row["subject_name"]

                try:
                    marks_dict = json.loads(row["marks_json"])
                except:
                    continue

                if subj not in stats:
                    stats[subj] = {}

                for head, mark in marks_dict.items():

                    try:
                        mark = float(mark)
                    except:
                        continue      # skip AB / NA / empty

                    if mark <= 0:
                        continue      # absent / detain

                    if head not in stats[subj]:
                        stats[subj][head] = {"marks_list": []}

                    stats[subj][head]["marks_list"].append(mark)

            # ----------------------
            # Calculate statistics
            # ----------------------

            for subj, heads in stats.items():

                for head, data in heads.items():

                    marks_list = data["marks_list"]

                    if not marks_list:
                        stats[subj][head] = {
                            "min": "-",
                            "max": "-",
                            "appeared": 0,
                            "passed": 0,
                            "pass_percent": 0,
                            "above60": 0,
                            "na": True
                        }
                        continue

                    appeared = len(marks_list)

                    min_marks = min(marks_list)
                    max_marks = max(marks_list)

                    pass_mark = max_marks * 0.4
                    above60_mark = max_marks * 0.6

                    passed = len([m for m in marks_list if m >= pass_mark])
                    above60 = len([m for m in marks_list if m >= above60_mark])

                    stats[subj][head] = {
                        "min": min_marks,
                        "max": max_marks,
                        "appeared": appeared,
                        "passed": passed,
                        "pass_percent": round((passed / appeared) * 100, 2) if appeared else 0,
                        "above60": round((above60 / appeared) * 100, 2) if appeared else 0,
                        "na": False
                    }

    cur.close()
    db.close()

    return render_template(
        "k7B.html",
        schemes=schemes,
        years=years,
        semesters=semesters,
        departments=departments,
        courses=courses,
        show_marks=show_marks,
        stats=stats,
        scheme_id=scheme_id,
        year=year,
        semester=semester,
        dept_id=dept_id,
        course_code=subject_name
    )
@app.route("/k7c")
def k7c():
    import json

    db = get_db()
    cur = db.cursor(dictionary=True)

    dept_id = session.get("dept_id")
    academic_year = request.args.get("academic_year")
    semester = request.args.get("semester")

    # ---- Years dropdown ----
    cur.execute("""
        SELECT DISTINCT academic_year
        FROM gazette_1
        WHERE dept_id=%s
        ORDER BY academic_year DESC
    """, (dept_id,))
    years = cur.fetchall()

    if not academic_year or not semester:
        return render_template(
            "k7C.html",
            years=years,
            academic_year=None,
            semester=None,
            data=[]
        )

    if semester not in ["5", "6"]:
        return "K7C report is only for Final Year Semester (5 or 6)"

    # ---- Subjects ----
    cur.execute("""
        SELECT course_code, course_title
        FROM courses
        WHERE semester=%s AND dept_id=%s
    """, (semester, dept_id))
    subjects = cur.fetchall()

    final_data = []
    sr = 1

    # ---- Fixed max marks ----
    max_marks = {"FA_TH": 50, "SA_TH": 100}

    for sub in subjects:
        subject_name = sub["course_title"]

        cur.execute("""
            SELECT marks_json
            FROM gazette_1
            WHERE subject_name=%s
            AND semester=%s
            AND academic_year=%s
            AND dept_id=%s
        """, (subject_name, semester, academic_year, dept_id))
        rows = cur.fetchall()

        total_fa = 0
        total_sa = 0
        student_count = 0

        for row in rows:
            marks_str = row.get("marks_json") or "{}"
            try:
                marks = json.loads(marks_str)  # convert string to dict
            except:
                marks = {}

            fa_marks = int(marks.get("FA_TH", 0))
            sa_marks = int(marks.get("SA_TH", 0))

            if fa_marks or sa_marks:
                total_fa += fa_marks
                total_sa += sa_marks
                student_count += 1

        fa_index = (total_fa * 100) / (student_count * max_marks["FA_TH"]) if student_count > 0 else 0
        sa_index = (total_sa * 100) / (student_count * max_marks["SA_TH"]) if student_count > 0 else 0

        difference = abs(fa_index - sa_index)
        remark = "YES" if difference > 20 else "NO"

        final_data.append({
            "sr": sr,
            "course_title": subject_name,
            "fa_index": round(fa_index, 2),
            "sa_index": round(sa_index, 2),
            "difference": round(difference, 2),
            "remark": remark
        })
        sr += 1

    cur.close()
    db.close()

    return render_template(
        "k7C.html",
        years=years,
        academic_year=academic_year,
        semester=semester,
        data=final_data
    )
@app.route("/api", methods=["GET", "POST"])
def api_page():

    db = get_db()
    cur = db.cursor(dictionary=True)

    # Departments fetch
    cur.execute("SELECT dept_id, dept_name FROM departments")
    departments = cur.fetchall()

    api = avg_api = fy_perf = sy_perf = ty_perf = None

    if request.method == "POST":

        academic_year = request.form["academic_year"]
        semester = request.form["semester"]

        role = session.get("role")
        session_dept = session.get("dept_id")

        dept_filter = ""
        params = [academic_year, semester]

        # Department Logic
        if role == "hod":
            dept_filter = " AND dept_id=%s "
            params.append(session_dept)

        elif role == "principal":
            selected_dept = request.form.get("dept_id")
            if selected_dept and selected_dept != "all":
                dept_filter = " AND dept_id=%s "
                params.append(selected_dept)

        # ------------------------
        # Z - Appeared Students
        # ------------------------
        cur.execute(f"""
            SELECT COUNT(DISTINCT enrollment_no) AS total_appeared
            FROM gazette_1
            WHERE academic_year=%s AND semester=%s
            {dept_filter}
        """, tuple(params))

        row = cur.fetchone()
        Z = row["total_appeared"] if row else 0

        # ------------------------
        # Y - Passed Students
        # ------------------------
        cur.execute(f"""
            SELECT COUNT(DISTINCT enrollment_no) AS total_passed
            FROM gazette_1
            WHERE academic_year=%s AND semester=%s
            AND RESULT='Pass'
            {dept_filter}
        """, tuple(params))

        row = cur.fetchone()
        Y = row["total_passed"] if row else 0

        # ------------------------
        # X - Mean Percentage
        # ------------------------
        cur.execute(f"""
            SELECT enrollment_no, SUM(TOTAL) AS total_marks
            FROM gazette_1
            WHERE academic_year=%s AND semester=%s
            AND RESULT='Pass'
            {dept_filter}
            GROUP BY enrollment_no
        """, tuple(params))

        rows = cur.fetchall()

        percentages = []
        for r in rows:
            total_marks = r["total_marks"]
            percentage = (total_marks / 600) * 100   # change 600 if needed
            percentages.append(percentage)

        X = sum(percentages)/len(percentages) if percentages else 0

        # ------------------------
        # API Calculation
        # ------------------------
        api = (float(X)/10) * (float(Y)/float(Z)) if Z != 0 else 0
        avg_api = api
        fy_perf = avg_api * 2.5
        sy_perf = avg_api * 2.0
        ty_perf = avg_api * 1.5

    cur.close()
    db.close()

    return render_template("k7F.html",
                           api=api,
                           avg_api=avg_api,
                           fy_perf=fy_perf,
                           sy_perf=sy_perf,
                           ty_perf=ty_perf,
                           departments=departments)
@app.route("/staff/dashboard")
def staff_dashboard():

    if session.get("role") != "staff":
        return redirect("/login")

    staff_id = session["user_id"]
    dept_id = session["dept_id"]

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT 
            c.course_code,
            c.course_title,
            c.year,
            c.semester
        FROM subject_allocation sa
        JOIN courses c ON sa.course_code = c.course_code
        WHERE sa.staff_id = %s AND sa.dept_id = %s
        ORDER BY c.year, c.semester
    """, (staff_id, dept_id))

    subjects = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        "staff_dashboard.html",
        staff_name=session["name"],
        dept_name=session["dept_name"],
        subjects=subjects
    )

@app.route("/mess_dashboard")
def mess_dashboard():
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT COUNT(*) AS total FROM girls_hostel_mess")
    total_girls = cur.fetchone()['total']

    cur.execute("""
        SELECT SUM(mess_fee) AS total 
        FROM girls_hostel_mess 
        WHERE status='Paid'
    """)
    total_paid = cur.fetchone()['total'] or 0

    cur.close()
    db.close()

    return render_template(
        "mess_dashboard.html",
        total_girls=total_girls,
        total_paid=total_paid
    )

@app.route("/add", methods=["GET","POST"])
def add():
    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT dept_id, dept_name FROM departments")
    departments = cur.fetchall()

    if request.method == "POST":
        data = request.form
        cur.execute("""
            INSERT INTO girls_hostel_mess
            (student_name, dept_id, diploma_year, month, year, mess_fee, status)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """,(
            data['student_name'],
            data['dept_id'],
            data['diploma_year'],
            data['month'],
            data['year'],
            data['mess_fee'],
            data['status']
        ))
        db.commit()
        cur.close()
        db.close()
        return redirect("/mess_dashboard")

    cur.close()
    db.close()
    return render_template("mess_add.html", departments=departments)

@app.route("/mess_view")
def mess_view():
    db = get_db()
    cur = db.cursor(dictionary=True)

    month = request.args.get("month")
    year = request.args.get("year")

    # 🔹 Dropdown data (ONLY existing values)
    cur.execute("SELECT DISTINCT month FROM girls_hostel_mess ORDER BY month")
    months = cur.fetchall()

    cur.execute("SELECT DISTINCT year FROM girls_hostel_mess ORDER BY year")
    years = cur.fetchall()

    # 🔹 Records
    query = """
        SELECT g.*, d.dept_name
        FROM girls_hostel_mess g
        JOIN departments d ON g.dept_id = d.dept_id
        WHERE 1
    """
    params = []

    if month:
        query += " AND g.month=%s"
        params.append(month)

    if year:
        query += " AND g.year=%s"
        params.append(year)

    cur.execute(query, params)
    records = cur.fetchall()

    cur.close()
    db.close()

    return render_template(
        "mess_view.html",
        records=records,
        months=months,
        years=years,
        sel_month=month,
        sel_year=year
    )
@app.route("/mess/edit/<int:id>", methods=["GET","POST"])
def edit_mess(id):
    db = get_db()
    cur = db.cursor(dictionary=True)

    # departments dropdown
    cur.execute("SELECT dept_id, dept_name FROM departments")
    departments = cur.fetchall()

    # existing record
    cur.execute("SELECT * FROM girls_hostel_mess WHERE id=%s", (id,))
    record = cur.fetchone()

    if request.method == "POST":
        data = request.form
        cur.execute("""
            UPDATE girls_hostel_mess SET
            student_name=%s,
            dept_id=%s,
            diploma_year=%s,
            month=%s,
            year=%s,
            mess_fee=%s,
            status=%s
            WHERE id=%s
        """,(
            data['student_name'],
            data['dept_id'],
            data['diploma_year'],
            data['month'],
            data['year'],
            data['mess_fee'],
            data['status'],
            id
        ))
        db.commit()
        cur.close()
        db.close()
        return redirect("/view")

    cur.close()
    db.close()
    return render_template(
        "mess_edit.html",
        record=record,
        departments=departments
    )
@app.route("/mess/delete/<int:id>")
def delete_mess(id):
    db = get_db()
    cur = db.cursor()

    cur.execute("DELETE FROM girls_hostel_mess WHERE id=%s", (id,))
    db.commit()

    cur.close()
    db.close()
    return redirect("/view")

@app.route("/gazette_add")
def gazette_add():

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT id, scheme_name FROM schemes")
    schemes = cur.fetchall()

    cur.close()
    db.close()

    return render_template("gazette_add.html", schemes=schemes)
app.secret_key = "supersecretkey"

@app.route("/upload_gazette", methods=["POST"])
def upload_gazette():

    import pandas as pd
    import json

    dept_id = session.get("dept_id")
    scheme_id = request.form.get("scheme_id")
    academic_year = request.form.get("academic_year")
    selected_semesters = request.form.getlist("semesters")
    file = request.files.get("file")

    if not file:
        flash("Please upload Excel file", "danger")
        return redirect(url_for("gazette_add"))

    if not selected_semesters:
        flash("Please select semester", "danger")
        return redirect(url_for("gazette_add"))

    db = get_db()
    cur = db.cursor()

    # -----------------------------
    # ✅ Read Excel (Multi Header)
    # -----------------------------
    excel_sheets = pd.read_excel(file, sheet_name=None, header=[0,1])

    # -----------------------------
    # ✅ Match Sheet by Semester Number
    # -----------------------------
    matched_sheets = []

    for sheet_name in excel_sheets.keys():
        for sem in selected_semesters:
            if sem in sheet_name:
                matched_sheets.append(sheet_name)

    if not matched_sheets:
        flash("Selected semester sheet not found in Excel", "danger")
        return redirect(url_for("gazette_add"))

    # -----------------------------
    # ✅ Get Scheme Subjects
    # -----------------------------
    cur.execute("SELECT course_title FROM courses WHERE scheme_id=%s", (scheme_id,))
    scheme_subjects = [row[0].strip().upper() for row in cur.fetchall()]

    if not scheme_subjects:
        flash("No subjects found for selected scheme", "danger")
        return redirect(url_for("gazette_add"))

    # =============================
    # 🔥 PROCESS EACH MATCHED SHEET
    # =============================
    for sheet in matched_sheets:

        df = excel_sheets[sheet]

        # -----------------------------
        # ✅ Extract Excel Subjects
        # -----------------------------
        excel_subjects = []

        for col in df.columns[2:]:
            subject = str(col[0]).strip().upper()
            if subject and subject not in excel_subjects:
                excel_subjects.append(subject)

        # -----------------------------
        # ✅ VALIDATION
        # -----------------------------
        missing_subjects = []

        for sub in scheme_subjects:
            if sub not in excel_subjects:
                missing_subjects.append(sub)

        if missing_subjects:
            flash("Check your sheet header as per scheme", "danger")
            return redirect(url_for("gazette_add"))

        # -----------------------------
        # ✅ INSERT DATA
        # -----------------------------
        for index, row in df.iterrows():

            enrollment = str(row[0]).strip()
            name = str(row[1]).strip()

            if enrollment == "" or enrollment.lower() == "nan":
                continue

            for subject in scheme_subjects:

                subject_total = 0
                marks_dict = {}

                for col in df.columns:

                    col_subject = str(col[0]).strip().upper()
                    col_type = str(col[1]).strip().upper()

                    if col_subject == subject:

                        mark_value = row[col]

                        if pd.notna(mark_value):

                            value = str(mark_value).strip()

    # ignore blank / space
                            if value == "" or value == "nan":
                                continue

                            try:
                                value_int = int(float(value))   # handles 10.0 also
                            except:
                                continue   # if still invalid skip

    # Only add if >0
                            if value_int > 0:
                                marks_dict[col_type] = value_int
                                subject_total += value_int

                if subject_total == 0:
                    result = ""
                elif subject_total >= 40:
                    result = "PASS"
                else:
                    result = "FAIL"
                cur.execute("""
                    INSERT INTO gazette_1
                    (academic_year, semester, enrollment_no,
                     student_name, subject_name,
                     marks_json, TOTAL, RESULT,
                     dept_id, scheme_id)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    academic_year,
                    selected_semesters[0],   # first selected semester
                    enrollment,
                    name,
                    subject,
                    json.dumps(marks_dict),
                    subject_total,
                    result,
                    dept_id,
                    scheme_id
                ))

    db.commit()
    cur.close()
    db.close()

    flash("Gazette Uploaded Successfully", "success")
    return redirect(url_for("gazette_view"))
@app.route("/gazette_view")
def gazette_view():

    academic_year = request.args.get("academic_year")
    semester = request.args.get("semester")
    enrollment_no = request.args.get("enrollment_no")
    scheme_id = request.args.get("scheme_id")

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("SELECT id, scheme_name FROM schemes")
    schemes = cur.fetchall()

    query = "SELECT * FROM gazette_1 WHERE 1=1"
    values = []

    if session.get("role") != "principal":
        query += " AND dept_id=%s"
        values.append(session.get("dept_id"))

    if academic_year:
        query += " AND academic_year=%s"
        values.append(academic_year)

    if semester:
        query += " AND semester=%s"
        values.append(semester)

    if enrollment_no:
        query += " AND enrollment_no=%s"
        values.append(enrollment_no)

    if scheme_id:
        query += " AND scheme_id=%s"
        values.append(scheme_id)

    query += " ORDER BY enrollment_no, subject_name"

    cur.execute(query, tuple(values))
    rows = cur.fetchall()

    import json

    heads = set()

    for row in rows:
        marks_json = row.get("marks_json")
        if marks_json:
            try:
                marks_dict = json.loads(marks_json)
            except:
                marks_dict = {}
            row["marks_json"] = marks_dict

            # Add keys to heads set (only strings)
            for key in marks_dict.keys():
                if isinstance(key, str):
                    heads.add(key)
        else:
            row["marks_json"] = {}

    heads = sorted(list(heads))  

    cur.close()
    db.close()

    students = {}

    for row in rows:
        enroll = row["enrollment_no"]

        if enroll not in students:
            students[enroll] = {
                "academic_year": row["academic_year"],
                "semester": row["semester"],
                "student_name": row["student_name"],
                "subjects": []
            }

        students[enroll]["subjects"].append(row)

    return render_template(
        "gazette_view.html",
        students=students,
        schemes=schemes,
        selected_scheme=scheme_id,
        heads=heads
    )

from openpyxl import Workbook
from openpyxl.styles import Alignment, Protection
from flask import send_file
import os

@app.route("/download_template/<int:scheme_id>")
def download_template(scheme_id):
    dept_id = session.get("dept_id")
    db = get_db()
    cur = db.cursor(dictionary=True)

    # Get mark types from scheme
    cur.execute("""
        SELECT head1, head2, head3, head4, head5
        FROM schemes
        WHERE id=%s
    """, (scheme_id,))
    scheme = cur.fetchone()
    if not scheme:
        return "No Scheme Found"

    mark_types = [v.strip() for v in scheme.values() if v and v.strip() != ""]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sem 1"

    # Protect sheet
    ws.protection.sheet = True
    ws.protection.enable()

    # Add Enrollment and Name columns
    ws["A1"] = "Enrollment"
    ws["B1"] = "Name"
    ws["A1"].protection = Protection(locked=True)
    ws["B1"].protection = Protection(locked=True)

    # Get subjects
    cur.execute("""
        SELECT course_title
        FROM courses
        WHERE dept_id=%s AND scheme_id=%s
    """, (dept_id, scheme_id))
    subjects = [row["course_title"].strip() for row in cur.fetchall()]

    col = 3
    for subject in subjects:
        # Merge cells for subject header
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1, end_column=col+len(mark_types)-1)
        ws.cell(row=1, column=col).value = subject
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=1, column=col).protection = Protection(locked=True)

        # Mark type headers
        for i, mt in enumerate(mark_types):
            cell = ws.cell(row=2, column=col+i)
            cell.value = mt
            cell.alignment = Alignment(horizontal="center")
            cell.protection = Protection(locked=True)

        col += len(mark_types)

    # Add sample student row (unlocked for input)
    ws.append(["2201350000", "Student Name"] + [0]*(len(subjects)*len(mark_types)))
    for c in range(3, col):
        ws.cell(row=3, column=c).protection = Protection(locked=True)

    file_path = "gazette_template.xlsx"
    wb.save(file_path)

    cur.close()
    db.close()
    return send_file(file_path, as_attachment=True)


@app.route('/register_scheme', methods=['GET', 'POST'])
def register_scheme():
    if request.method == 'POST':

        scheme_name = request.form['scheme_name']
        year_start = request.form['year_start']
        head1 = request.form['head1']
        head2 = request.form['head2']
        head3 = request.form['head3']
        head4 = request.form['head4']
        head5 = request.form['head5']

        db = get_db()
        cur = db.cursor(dictionary=True)

        cur.execute("""
            INSERT INTO schemes 
            (scheme_name, year_start, head1, head2, head3, head4, head5)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (scheme_name, year_start, head1, head2, head3, head4, head5))

        db.commit()   # ✅ VERY IMPORTANT

        cur.close()
        db.close()

        flash("Scheme Registered Successfully!")
        return redirect('/register_scheme')

    return render_template('scheme_reg.html')

@app.route('/view_schemes')
def view_schemes():

    db = get_db()
    cur = db.cursor(dictionary=True)
    
    cur.execute("SELECT * FROM schemes ORDER BY id DESC")
    schemes = cur.fetchall()

    print(schemes)  # Debug

    cur.close()
    db.close()

    return render_template('view_scheme.html', schemes=schemes)

from flask import jsonify

@app.route("/get_scheme_heads/<int:scheme_id>")
def get_scheme_heads(scheme_id):

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT head1, head2, head3, head4, head5
        FROM schemes
        WHERE id=%s
    """, (scheme_id,))

    scheme = cur.fetchone()

    cur.close()
    db.close()

    return jsonify(scheme)

# @app.route("/task")
# # def task():
#     return render_template("task.html")

# @app.route("/SD")
# def SD():
#     return render_template("SD.html")

@app.route("/staff_dash")
def staff_dash():

    db = get_db()
    cur = db.cursor(dictionary=True)

    # Get all departments
    cur.execute("SELECT * FROM departments")
    departments = cur.fetchall()

    return render_template(
        "staff_dash.html",
        departments=departments
    )


# ---------------- SHOW TASKS ----------------

@app.route("/staff_tasks/<int:dept_id>")
def staff_tasks(dept_id):

    db = get_db()
    cur = db.cursor(dictionary=True)

    cur.execute("""
        SELECT id,title,description,status
        FROM tasks
        WHERE dept_id=%s
    """, (dept_id,))

    tasks = cur.fetchall()

    return render_template("task.html", tasks=tasks)


# ---------------- ASSIGN TASK ----------------

@app.route("/assign_task", methods=["GET", "POST"])
def assign_task():

    db = get_db()
    cur = db.cursor(dictionary=True)

    if request.method == "POST":

        staff_id = request.form["staff_id"]
        dept_id = request.form["dept_id"]
        title = request.form["title"]
        description = request.form["description"]

        cur.execute("""
            INSERT INTO tasks
            (staff_id, dept_id, title, description, assigned_date)
            VALUES (%s, %s, %s, %s, CURDATE())
        """, (staff_id, dept_id, title, description))

        db.commit()

    # Staff list
    cur.execute("SELECT * FROM staff")
    staff_list = cur.fetchall()

    # Departments
    cur.execute("SELECT * FROM departments")
    departments = cur.fetchall()

    return render_template(
        "assign_task.html",
        staff_list=staff_list,
        departments=departments
    )

# ================= COMPLETE TASK =================

@app.route("/mark_complete/<int:task_id>", methods=["POST"])
def mark_complete(task_id):

    db = get_db()
    cur = db.cursor()

    cur.execute("""
        UPDATE tasks
        SET status='Completed'
        WHERE id=%s
    """, (task_id,))

    db.commit()

    return redirect(request.referrer)

# ================= RUN =================
if __name__ == "__main__":
    import os
    # Render कडून पोर्ट मिळवण्यासाठी, नसेल तर १०००० वापरण्यासाठी
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
